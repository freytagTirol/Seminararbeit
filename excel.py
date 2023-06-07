import os.path
import openpyxl
import csv
import logging

from globals import PATH


class Excel:
    def __init__(self, table_path, table):
        # Open excel table
        # xlsx_file = PATH + "df_litigation_MDA.xlsx"
        xlsx_file = PATH + "Regression_Variables.xlsx"
        self.wb_obj = openpyxl.load_workbook(xlsx_file)
        self.sheet = self.wb_obj.active
        self.table_path = table_path
        self.table = table

    def get_litigation_for_id(self, mda_id):
        for row in self.sheet.iter_cols(min_col=1, max_col=1, min_row=2):
            for cell in row:
                # print(cell.value)
                if cell.value == int(mda_id):
                    # print(cell.value)
                    # print(self.sheet.cell(row=cell.row, column=24).value)
                    # print(f'{cell.value}.txt has {self.sheet.cell(row=cell.row, column=24).value} litigation')
                    return self.sheet.cell(row=cell.row, column=24).value

    # function to export the data to a csv file

    def export_to_csv(self):
        chat_gpt_estimations = self.table
        table_path = self.table_path + '/table.csv'
        try:
            # Extracting the fieldnames from the first dictionary in the list
            fieldnames = chat_gpt_estimations[0].keys()
        except Exception as e:
            print("ERROR: chat_gpt_estimations is broken or empty")
            logging.exception(e)
            exit(1)

            # Writing the data to the CSV file
        with open(table_path, 'w', newline='') as file:
            writer = csv.DictWriter(file, fieldnames=fieldnames)
            writer.writeheader()
            writer.writerows(chat_gpt_estimations)

        print("Tabelle: " + str(chat_gpt_estimations))
        print("Tabelle erfolgreich gespeichert:" + table_path)

    def export_to_xlsx(self):
        """
        Exports the table to csv
        long_prompt_used = -1 means mda_file was not sent to OpenAI
        long_prompt_used = 0 means mda_file was short enough
        long_prompt_used = 1 means mda_file was cut off at split_token_limit
        long_prompt_used = 2 means mda_file was split and summarized

        :return:
        """
        chat_gpt_estimations = self.table
        full_table_path = self.table_path + "/table.xlsx"
        try:
            chat_gpt_estimations[0].keys()
        except Exception as e:
            print("ERROR: chat_gpt_estimations is broken or empty")
            logging.exception(e)
            exit(1)

        # check if there is a res table already
        if not os.path.isfile(full_table_path):
            # create new xlsx (wb)
            wb = openpyxl.Workbook()

            ws = wb.active
            for column in self.sheet:
                for cell in column:
                    ws[cell.coordinate].value = cell.value

            # add columns GPT_lit GPT_prob GPT_tokens_used
            # ws.insert_cols(ws.max_column + 1, amount=3)

            # add column names
            ws.cell(row=1, column=ws.max_column + 1).value = "GPT_lit"
            ws.cell(row=1, column=ws.max_column + 1).value = "GPT_prob"
            ws.cell(row=1, column=ws.max_column + 1).value = "GPT_tokens_used"
            ws.cell(row=1, column=ws.max_column + 1).value = "GPT_long_prompt"
        else:
            wb = openpyxl.load_workbook(full_table_path)
            ws = wb.active
            # TODO: Add check if row 1 headlines are correct in ws

        for entry in chat_gpt_estimations:
            for row in ws.iter_cols(min_col=1, max_col=1, min_row=2):
                for cell in row:
                    if cell.value == int(entry["id"]):
                        ws.cell(row=cell.row, column=ws.max_column - 3).value = entry["litigation"]
                        ws.cell(row=cell.row, column=ws.max_column - 2).value = entry["probability"]
                        ws.cell(row=cell.row, column=ws.max_column - 1).value = entry["tokens_used"]
                        ws.cell(row=cell.row, column=ws.max_column).value = entry["long_prompt_used"]

        wb.save(full_table_path)

        print("Tabelle: " + str(chat_gpt_estimations))
        print("Tabelle erfolgreich gespeichert: " + full_table_path)

    @staticmethod
    def get_data(xlsx_path, col_name):
        wb = openpyxl.load_workbook(xlsx_path)
        ws = wb.active

        # get all rows in col_name
        for col in ws.iter_cols(min_row=1, max_row=1):
            if col[0].value == col_name:
                rows = []
                for row in ws.iter_rows(min_row=2, min_col=col[0].column, max_col=col[0].column):
                    for cell in row:
                        rows.append(cell.value)
                return rows
